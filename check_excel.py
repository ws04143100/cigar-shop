import openpyxl
import os
import glob

# Find the latest wholesale file
files = glob.glob('assets/*批發表*.xlsx')
print(f"Files found: {files}")

if files:
    latest = sorted(files, reverse=True)[0]
    print(f"Reading: {latest}")
    wb = openpyxl.load_workbook(latest)
    ws = wb.active
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    print("Row 1 (header):", [str(c.value) if c.value else "" for c in ws[1]])
    print("Row 2:", [str(c.value) if c.value else "" for c in ws[2]])
    print("Row 3:", [str(c.value) if c.value else "" for c in ws[3]])
